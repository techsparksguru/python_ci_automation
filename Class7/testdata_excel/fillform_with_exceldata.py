import time
from selenium import webdriver
from itertools import zip_longest
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support.ui import Select

wb = load_workbook(filename='TestExcel.xlsx')
ws = wb.active
lst = []

browser = webdriver.Chrome()
browser.get('http://www.practiceselenium.com/practice-form.html')
header_row = None
for i, row in enumerate(ws.iter_rows()):
    if header_row == None:
        header_row = [ c.internal_value for c in row ]
        continue
    row = dict(zip_longest(header_row, [ c.internal_value for c in row ]))
    print(row)
    lst.append(row)

for ele in lst:
    browser.find_element_by_name("firstname").send_keys(ele['firstname'])
    browser.find_element_by_name("lastname").send_keys(ele['lastname'])
    sex_element = next(element for element in browser.find_elements_by_name("sex") if element.get_attribute("value")==ele['sex'])
    sex_element.click()
    yrs_element = next(element for element in browser.find_elements_by_tag_name("input") if element.get_attribute("value")==str(ele['yrs']))
    yrs_element.click()
    browser.find_element_by_id('datepicker').send_keys(str(ele['date_stopped']))
    tea_element = next(element for element in browser.find_elements_by_tag_name("input"))
    tea_element.click()
    excited_element = next(element for element in browser.find_elements_by_name("tool"))
    excited_element.click()
    continents_select = Select(browser.find_element_by_id('continents'))
    continents_element = next(element for element in continents_select.options)
    continents_element.click()

    another_select_list = Select(browser.find_element_by_id('selenium_commands'))
    another_select_element = next(element for element in another_select_list.options)
    another_select_element.click()
    browser.refresh()

browser.quit()