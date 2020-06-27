from openpyxl import Workbook,load_workbook

wb = load_workbook(filename='TestExcel.xlsx')
ws = wb.active
print("Printing all cell values by iterating rows....\n")
for row in ws.iter_rows():
    for cell in row:
        print(str(cell.value) + "\t")
print("\n")