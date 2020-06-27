import time
from openpyxl import Workbook

wb = Workbook()
ws = wb.active  # active worksheet by default
print("Active Sheet title: " + ws.title)
ws1 = wb.create_sheet('testsheet', 0)
print("Created worksheet with title: " + ws1.title)
print("Changing the title....")
ws1.title = "sheet_title"
print("New title:" + ws1.title)
print("All sheets: " + str(wb.sheetnames))
print("highest row:" + str(ws1.max_row))
print("highest column:" + str(ws1.max_row))

ws1['A1'] = "Name"
ws1['B1'].value = "Website"
ws1['C1'].value = "address"

ws1['A2'].value = "Pradeep"
ws1['B2'].value = "www.seleniumframework.com"
ws1['C2'].value = "Europe"

wb.save("TestExcel.xlsx")
time.sleep(5)