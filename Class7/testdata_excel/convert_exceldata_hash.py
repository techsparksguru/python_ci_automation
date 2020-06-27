from itertools import zip_longest
from openpyxl import Workbook, load_workbook

wb = load_workbook(filename='TestExcel.xlsx')
ws = wb.active
#http://www.goalbrecht.com/2014/04/python-iterating-an-excel-worksheet/
header_row = None
for i, row in enumerate(ws.iter_rows()):
    if header_row == None:
        header_row = [c.internal_value for c in row]
        continue
    row = dict(zip_longest(header_row, [c.internal_value for c in row]))
    print(row)